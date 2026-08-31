# -*- coding: utf-8 -*-
"""Excel 读写工具 — 广告花费报表的模板同步、日期列管理、增量写入。

从 `ExpenseVerification/广告查询/ziniao_playwright_http_py3_v2.py` 迁移，
去掉了脚本里的硬编码路径（由 config 传入）。
"""

import glob as glob_mod
import os
import shutil
import threading
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook

# 月度 Excel 文件的并发写锁
EXCEL_LOCK = threading.Lock()

# 输出文件命名
TEMPLATE_BASENAME = "各站点产品日广告花费数据表_USD.xlsx"


def read_excel_to_nested_dict(file_path: str) -> dict:
    """读取 Excel，返回 sheet -> 站点 -> 关键词列表 的嵌套字典。

    每个 sheet 取前两列（站点、关键词），前向填充合并单元格，
    按站点分组收集关键词列表。
    """
    all_data: dict = {}
    excel_data = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")

    for sheet_name, df in excel_data.items():
        if sheet_name.startswith("~$"):
            continue
        if df.shape[1] < 2:
            continue

        df_processed = df.iloc[:, :2].copy()
        df_processed.columns = ["站点", "关键词"]
        df_processed["站点"] = df_processed["站点"].ffill()
        df_processed = df_processed.dropna(subset=["关键词"])
        df_processed["站点"] = df_processed["站点"].astype(str).str.strip()
        df_processed["关键词"] = df_processed["关键词"].astype(str).str.strip()
        df_processed = df_processed[df_processed["站点"] != "nan"]

        if not df_processed.empty:
            sheet_dict = {}
            for site, group in df_processed.groupby("站点"):
                sheet_dict[site] = group["关键词"].tolist()
            all_data[sheet_name] = sheet_dict

    return all_data


def parse_date_from_header(value):
    """解析 Excel 表头中的日期，统一返回 datetime.date 或 None。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ["%Y/%m/%d", "%Y-%m-%d"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        parts = value.replace("/", "-").split("-")
        if len(parts) == 3:
            try:
                return date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, TypeError):
                pass
    return None


def get_sheet_date_columns(ws) -> dict:
    """获取工作表表头行（第 1 行）中所有日期列及其列号。"""
    date_cols = {}
    for col in range(3, ws.max_column + 1):
        d = parse_date_from_header(ws.cell(row=1, column=col).value)
        if d is not None:
            date_cols[d] = col
    return date_cols


def get_completed_dates_per_sheet(file_path: str) -> dict:
    """读取月度 Excel，返回每个 sheet 已存在的日期集合。"""
    if not os.path.exists(file_path):
        return {}
    wb = load_workbook(file_path)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result[sheet_name] = set(get_sheet_date_columns(ws).keys())
    wb.close()
    return result


def get_completed_dates_from_output(dir_path: str) -> set:
    """扫描输出目录，返回已「完整写入」的日期集合。

    判定标准：所有 sheet 都有该日期列，且该列下所有有效商品行（B 列非空且非 '/'）
    都有数据（None 视为未写，'' 视为写过但无花费）。
    """
    completed: set = set()
    pattern = os.path.join(dir_path, f"{TEMPLATE_BASENAME.split('.')[0]}_*.xlsx")
    for f in glob_mod.glob(pattern):
        basename = os.path.basename(f)
        if "测试" in basename or basename == TEMPLATE_BASENAME:
            continue
        try:
            wb = load_workbook(f)
            all_sheet_complete_dates = None
            for sn in wb.sheetnames:
                ws = wb[sn]
                date_cols = get_sheet_date_columns(ws)
                complete_in_sheet = set()
                for d, col in date_cols.items():
                    all_filled = True
                    for row in range(2, ws.max_row + 1):
                        b_val = ws.cell(row=row, column=2).value
                        if b_val and str(b_val).strip() and str(b_val).strip() != "/":
                            if ws.cell(row=row, column=col).value is None:
                                all_filled = False
                                break
                    if all_filled:
                        complete_in_sheet.add(d)
                if all_sheet_complete_dates is None:
                    all_sheet_complete_dates = complete_in_sheet
                else:
                    all_sheet_complete_dates = all_sheet_complete_dates.intersection(complete_in_sheet)
            if all_sheet_complete_dates:
                completed.update(all_sheet_complete_dates)
            wb.close()
        except Exception:
            continue
    return completed


def map_site_results_to_operators(site_date_results: dict, original_result: dict) -> dict:
    """将 (国家, 日期) → 商品 的查询结果映射回 operator → 国家 → 日期 → 商品。"""
    operator_results = {}
    for operator, sites in original_result.items():
        operator_results[operator] = {}
        for site, products in sites.items():
            operator_results[operator][site] = {}
            for (s, d), spend_dict in site_date_results.items():
                if s == site:
                    operator_results[operator][site][d] = {}
                    for product in products:
                        if product and product != "/":
                            operator_results[operator][site][d][product] = spend_dict.get(product.strip())
    return operator_results


def write_result_to_excel(result_dict: dict, dir_path: str, query_date) -> None:
    """将单日查询结果写入月度 Excel，按日期顺序插入/更新列。

    逻辑：
      1. 目标月度文件不存在则从模板复制一份；
      2. 每 sheet 若该日期列已存在则更新，否则按时间顺序插入新列；
      3. 结构损坏（列数 < 3）则删除重建；
      4. 线程安全：EXCEL_LOCK 保护并发写入。

    :param result_dict: 操作人 -> 站点 -> 商品 的嵌套字典（仅一个日期的数据）
    :param dir_path: 输出目录（模板文件所在目录）
    :param query_date: 查询日期（datetime.date）
    """
    with EXCEL_LOCK:
        date_str = f"{query_date.year}/{query_date.month}/{query_date.day}"
        file_str = f"{query_date.year}_{query_date.month}"

        origin_file = os.path.join(dir_path, TEMPLATE_BASENAME)
        set_file = os.path.join(dir_path, f"各站点产品日广告花费数据表_USD_{file_str}.xlsx")

        # 确保输出文件正确初始化
        need_reinit = False
        if not os.path.exists(set_file):
            need_reinit = True
        else:
            try:
                wb_check = load_workbook(set_file)
                for sn in wb_check.sheetnames:
                    if wb_check[sn].max_column < 3:
                        need_reinit = True
                        break
                wb_check.close()
            except Exception:
                need_reinit = True

        if need_reinit:
            if not os.path.exists(origin_file):
                raise FileNotFoundError(f"原始模板文件不存在：{origin_file}")
            shutil.copy2(origin_file, set_file)

        wb = load_workbook(set_file)
        total_written = 0

        for operator_name, sites_data in result_dict.items():
            if operator_name not in wb.sheetnames:
                continue
            ws = wb[operator_name]
            date_cols = get_sheet_date_columns(ws)

            # 确定目标列号
            if query_date in date_cols:
                target_col = date_cols[query_date]
            else:
                insert_pos = None
                sorted_existing = sorted(date_cols.keys())
                for existing_date in sorted_existing:
                    if existing_date > query_date:
                        insert_pos = date_cols[existing_date]
                        break
                if insert_pos is None:
                    last_data_col = 2
                    for col in range(3, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value is not None:
                            last_data_col = col
                    target_col = last_data_col + 1
                    ws.cell(row=1, column=target_col, value=date_str)
                else:
                    target_col = insert_pos
                    ws.insert_cols(target_col)
                    ws.cell(row=1, column=target_col, value=date_str)

            # 构建 A 列 行号 -> 站点名称 映射
            row_to_site = {}
            current_site = None
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_col == 1 and merged_range.max_col == 1:
                    site_value = ws.cell(row=merged_range.min_row, column=1).value
                    if site_value:
                        site_value = str(site_value).strip()
                        for row in range(merged_range.min_row, merged_range.max_row + 1):
                            row_to_site[row] = site_value
            for row in range(2, ws.max_row + 1):
                if row in row_to_site:
                    current_site = row_to_site[row]
                else:
                    cell_val = ws.cell(row=row, column=1).value
                    if cell_val and str(cell_val).strip():
                        current_site = str(cell_val).strip()
                    if current_site:
                        row_to_site[row] = current_site

            # 遍历站点数据，匹配写入
            sheet_written = 0
            for site_name, products_data in sites_data.items():
                if not products_data:
                    continue
                site_display_name = site_name.replace("站", "").strip()
                if site_name == "美国本土店铺":
                    site_display_name = "美国"

                for row in range(2, ws.max_row + 1):
                    cell_site = row_to_site.get(row, "")
                    if not cell_site:
                        continue
                    site_match = (site_name == cell_site) or (site_display_name in cell_site)
                    if not site_match:
                        continue
                    cell_keyword = ws.cell(row=row, column=2).value
                    if not cell_keyword:
                        continue
                    keyword = str(cell_keyword).strip()
                    if keyword in products_data:
                        fill_val = products_data[keyword]
                        ws.cell(row=row, column=target_col,
                                value=fill_val if fill_val is not None else "")
                        sheet_written += 1

            total_written += sheet_written

        wb.save(set_file)
        wb.close()
