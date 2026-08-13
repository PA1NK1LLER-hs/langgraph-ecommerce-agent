"""
# 适用环境python3
"""
import datetime
import os
import platform
import shutil
import time
import traceback
import uuid
import json
from concurrent.futures import ThreadPoolExecutor
from typing import Literal

import pandas as pd
import requests
import subprocess

from openpyxl import load_workbook
from playwright import sync_api
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

""" 需要从系统角标将紫鸟浏览器完全退出后再运行"""
is_windows = platform.system() == 'Windows'
is_mac = platform.system() == 'Darwin'
is_linux = platform.system() == 'Linux'

# todo 1、修改client_path：紫鸟客户端在本设备的路径
if is_windows:
    client_path = os.getenv("ZINIAO_CLIENT_PATH", R'D:\software\ziniao\ziniao.exe')  # 紫鸟客户端在本设备的路径，V5程序名为starter.exe，V6程序名为ziniao.exe
elif is_linux:
    client_path = os.getenv("ZINIAO_CLIENT_PATH", R'/opt/ziniao/ziniaobrowser')  # 紫鸟客户端在本设备的路径
else:
    client_path = os.getenv("ZINIAO_CLIENT_PATH", R'ziniao')  # 客户端程序名称
socket_port = 16851  # 系统未被占用的端口

_ziniao_company = os.getenv("ZINIAO_COMPANY")
_ziniao_username = os.getenv("ZINIAO_USERNAME")

def _get_user_info():
    """延迟校验：仅在 RPA 工具实际调用时才要求环境变量已设置。"""
    if not _ziniao_company:
        raise RuntimeError("ZINIAO_COMPANY 未设置，紫鸟 RPA 无法运行")
    if not _ziniao_username:
        raise RuntimeError("ZINIAO_USERNAME 未设置，紫鸟 RPA 无法运行")
    return {
        "company": _ziniao_company,
        "username": _ziniao_username,
        "password": os.getenv("ZINIAO_PASSWORD", ""),
    }

user_info = None  # 延迟初始化，由 _get_user_info() 在调用时校验

def read_excel_to_nested_dict(file_path):
    """
    读取Excel文件，返回 sheet -> 站点 -> 关键词列表 的嵌套字典结构

    参数:
        file_path: Excel文件路径

    返回:
        dict: {
            'sheet页名称': {
                '站点名称': ['关键词1', '关键词2', ...],
                ...
            },
            ...
        }
    """
    all_data = {}

    # 读取所有sheet页
    excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')

    # 遍历每个sheet页
    for sheet_name, df in excel_data.items():
        # 过滤临时文件
        if sheet_name.startswith('~$'):
            continue

        # 检查必要的列是否存在
        if df.shape[1] < 2:
            continue

        # 获取前两列（站点和关键词）
        df_processed = df.iloc[:, :2].copy()
        df_processed.columns = ['站点', '关键词']

        # 前向填充站点列（处理合并单元格）
        df_processed['站点'] = df_processed['站点'].ffill()

        # 删除关键词为空的行
        df_processed = df_processed.dropna(subset=['关键词'])

        # 转换为字符串并去除空格
        df_processed['站点'] = df_processed['站点'].astype(str).str.strip()
        df_processed['关键词'] = df_processed['关键词'].astype(str).str.strip()

        # 删除空站点
        df_processed = df_processed[df_processed['站点'] != 'nan']

        if not df_processed.empty:
            # 按站点分组，构建嵌套字典
            sheet_dict = {}
            for site, group in df_processed.groupby('站点'):
                # 将同一站点下的所有关键词转为列表
                keywords = group['关键词'].tolist()
                sheet_dict[site] = keywords

            all_data[sheet_name] = sheet_dict

    return all_data
def kill_process(version: Literal["v5", "v6"]):
    """
    杀紫鸟客户端进程
    :param version: 客户端版本
    """
    # 确认是否继续
    confirmation = print("即将关闭紫鸟浏览器的主进程")

    if is_windows:
        if version == "v5":
            process_name = 'SuperBrowser.exe'
        else:
            process_name = 'ziniao.exe'
        os.system('taskkill /f /t /im ' + process_name)
        time.sleep(3)



def start_browser():
    """
    启动客户端
    :return:
    """
    try:
        if is_windows:
            cmd = [client_path, '--run_type=web_driver', '--ipc_type=http', '--port=' + str(socket_port)]
        elif is_mac:
            cmd = ['open', '-a', client_path, '--args', '--run_type=web_driver', '--ipc_type=http',
                   '--port=' + str(socket_port)]
        elif is_linux:
            cmd = [client_path, '--no-sandbox', '--run_type=web_driver', '--ipc_type=http', '--port=' + str(socket_port)]
        else:
            exit()
        subprocess.Popen(cmd)
        time.sleep(5)
    except Exception:
        print('start browser process failed: ' + traceback.format_exc())
        exit()


def update_core():
    """
    下载所有内核，打开店铺前调用，需客户端版本5.285.7以上
    因为http有超时时间，所以这个action适合循环调用，直到返回成功
    """
    data = {
        "action": "updateCore",
        "requestId": str(uuid.uuid4()),
    }
    data.update(_get_user_info())
    while True:
        result = send_http(data)
        print(result)
        if result is None:
            print("等待客户端启动...")
            time.sleep(2)
            continue
        if result.get("statusCode") is None or result.get("statusCode") == -10003:
            print("当前版本不支持此接口，请升级客户端")
            return
        elif result.get("statusCode") == 0:
            print("更新内核完成")
            return
        else:
            print(f"等待更新内核: {json.dumps(result)}")
            time.sleep(2)


def send_http(data):
    """
    通讯方式
    :param data:
    :return:
    """
    try:
        url = 'http://127.0.0.1:{}'.format(socket_port)
        response = requests.post(url, json.dumps(data).encode('utf-8'), timeout=120)
        return json.loads(response.text)
    except Exception as err:
        print(err)


def delete_all_cache():
    """
    删除所有店铺缓存
    非必要的，如果店铺特别多、硬盘空间不够了才要删除
    """
    if not is_windows:
        return
    local_appdata = os.getenv('LOCALAPPDATA')
    cache_path = os.path.join(local_appdata, 'SuperBrowser')
    if os.path.exists(cache_path):
        shutil.rmtree(cache_path)


def delete_all_cache_with_path(path):
    """
    :param path: 启动客户端参数使用--enforce-cache-path时设置的缓存路径
    删除所有店铺缓存
    非必要的，如果店铺特别多、硬盘空间不够了才要删除
    """
    if not is_windows:
        return
    cache_path = os.path.join(path, 'SuperBrowser')
    if os.path.exists(cache_path):
        shutil.rmtree(cache_path)


def open_store(store_info, isWebDriverReadOnlyMode=0, isprivacy=0, isHeadless=0, cookieTypeSave=0, jsInfo=""):
    request_id = str(uuid.uuid4())
    data = {
        "action": "startBrowser"
        , "isWaitPluginUpdate": 0
        , "isHeadless": isHeadless
        , "requestId": request_id
        , "isWebDriverReadOnlyMode": isWebDriverReadOnlyMode
        , "cookieTypeLoad": 0
        , "cookieTypeSave": cookieTypeSave
        , "runMode": "1"
        , "isLoadUserPlugin": False
        , "pluginIdType": 1
        , "privacyMode": isprivacy
    }
    data.update(_get_user_info())

    if store_info.isdigit():
        data["browserId"] = store_info
    else:
        data["browserOauth"] = store_info

    if len(str(jsInfo)) > 2:
        data["injectJsInfo"] = json.dumps(jsInfo)

    r = send_http(data)
    if str(r.get("statusCode")) == "0":
        return r
    elif str(r.get("statusCode")) == "-10003":
        print(f"login Err {json.dumps(r, ensure_ascii=False)}")
        exit()
    else:
        print(f"Fail {json.dumps(r, ensure_ascii=False)} ")
        exit()


def close_store(browser_oauth):
    request_id = str(uuid.uuid4())
    data = {
        "action": "stopBrowser"
        , "requestId": request_id
        , "duplicate": 0
        , "browserOauth": browser_oauth
    }
    data.update(_get_user_info())

    r = send_http(data)
    if str(r.get("statusCode")) == "0":
        return r
    elif str(r.get("statusCode")) == "-10003":
        print(f"login Err {json.dumps(r, ensure_ascii=False)}")
        exit()
    else:
        print(f"Fail {json.dumps(r, ensure_ascii=False)} ")
        exit()


def get_browser_list() -> list:
    request_id = str(uuid.uuid4())
    data = {
        "action": "getBrowserList",
        "requestId": request_id
    }
    data.update(_get_user_info())

    r = send_http(data)
    if str(r.get("statusCode")) == "0":
        print(r)
        return r.get("browserList")
    elif str(r.get("statusCode")) == "-10003":
        print(f"login Err {json.dumps(r, ensure_ascii=False)}")
        exit()
    else:
        print(f"Fail {json.dumps(r, ensure_ascii=False)} ")
        exit()


# 获取playwright浏览器会话
def get_browser_context(playwright, port):
    browser = playwright.chromium.connect_over_cdp("http://127.0.0.1:" + str(port))
    context = browser.contexts[0]
    return context


def open_ip_check(browser_context, ip_check_url):
    """
    打开ip检测页检测ip是否正常
    :param browser_context: playwright浏览器会话
    :param ip_check_url ip检测页地址
    :return 检测结果
    """
    try:
        page = browser_context.pages[0]
        page.goto(ip_check_url)
        success_button = page.locator('//button[contains(@class, "styles_btn--success")]')
        success_button.wait_for(timeout=60000)  # 等待查找元素60秒
        print("ip检测成功")
        return True
    except PlaywrightTimeoutError:
        print("ip检测超时")
        return False
    except Exception as e:
        print("ip检测异常:" + traceback.format_exc())
        return False


def open_launcher_page(browser_context, launcher_page):
    page = browser_context.pages[0]
    page.goto(launcher_page)

    page.wait_for_load_state(state="load")
    check_verified_status(page)


def get_exit():
    """
    关闭客户端
    :return:
    """
    data = {"action": "exit", "requestId": str(uuid.uuid4())}

    data.update(_get_user_info())

    print('@@ get_exit...')
    send_http(data)


def use_one_browser_run_task(browser):
    """
    打开一个店铺运行脚本
    :param browser: 店铺信息
    """
    # 如果要指定店铺ID, 获取方法:登录紫鸟客户端->账号管理->选择对应的店铺账号->点击"查看账号"进入账号详情页->账号名称后面的ID即为店铺ID
    store_id = browser.get('browserOauth')
    store_name = browser.get("browserName")
    result = browser.get("result_dict")
    # 打开店铺
    print(f"=====打开店铺：{store_name}=====")
    ret_json = open_store(store_id)
    print(ret_json)
    store_id = ret_json.get("browserOauth")
    if store_id is None:
        store_id = ret_json.get("browserId")
    product_dict= {}
    # 获取playwright浏览器会话
    with (sync_api.sync_playwright() as playwright):
        try:
            browser_context = get_browser_context(playwright, ret_json.get('debuggingPort'))
            if browser_context is None:
                print(f"=====关闭店铺：{store_name}=====")
                close_store(store_id)
                return

            # 获取ip检测页地址
            ip_check_url = ret_json.get("ipDetectionPage")
            if not ip_check_url:
                print("ip检测页地址为空，请升级紫鸟浏览器到最新版")
                print(f"=====关闭店铺：{store_name}=====")
                close_store(store_id)
                exit()
            ip_usable = open_ip_check(browser_context, ip_check_url)
            if ip_usable:
                print("ip检测通过，打开店铺平台主页")
                # 打开店铺平台主页后进行后续自动化操作
                # todo 后续的自动化操作

                open_launcher_page(browser_context, ret_json.get("launcherPage"))

                page = browser_context.new_page()
                # page.pause()
                page.goto("https://advertising.amazon.com/campaign-manager", timeout=60000)
                body = page.locator("//div[@id='globalBetaAllCampaigns:table']")
                body.wait_for(state= "visible",timeout=60000)


                for operator in result.keys():
                    states_product_dict = result[operator]
                    print(f"\n{'=' * 50}\n开始处理负责人: {operator}\n{'=' * 50}")

                    product_dict[operator] = {}
                    for key, value in states_product_dict.items():



                        print(f"\n=====查询站点: {key} =====")

                        product_dict[operator][key] = {}

                        delete_btn = page.get_by_role("button", name="删除所有")
                        if delete_btn.is_visible():
                            delete_btn.click()

                        today_option = page.get_by_role("option", name="昨天")
                        while today_option.count() <= 0:
                            try:
                                page.locator(
                                    "[id=\"UCM-CM-APP:globalBetaAllCampaigns:dateRangeFilter:openContainer\"]").click(force= True,timeout=60000)
                                page.get_by_role("option", name="昨天").click(force= True,timeout=5000)
                                break
                            except:
                                today_option = page.get_by_role("option", name="昨天")
                                page.reload()
                                body.wait_for(state="visible", timeout=60000)
                                pass
                        page.get_by_role("button", name="筛选条件").click()
                        page.get_by_role("option", name="国家/地区").click()

                        if key == "美国本土店铺":
                            search_text = "美国"
                        elif key == "澳洲站":
                            search_text = "澳大利亚"
                        else:
                            search_text = key.replace('站', '')

                        page.locator("label").filter(has_text=search_text).click()
                        page.get_by_role("button", name="应用").click(force= True)
                        for product in value:
                            if product == '/':
                                continue

                            search_box = page.get_by_placeholder("查找广告活动")
                            search_box.fill(product)
                            search_box.press("Enter")
                            page.wait_for_timeout(1000)

                            page.locator("[id=\"UCM-CM-APP:globalBetaAllCampaigns:overlay:loading\"] h4").filter(has_text="正在加载").wait_for(state = "hidden",timeout=60000)

                            page.locator(f"//span[@class='cell-renderer-content-text' and contains(text(), '{product}')]").first.wait_for(state = "visible",timeout=60000)


                            cell_renderer =  page.locator("//span[@class='cell-renderer-summary-text' and contains(@title, 'US')]")

                            # 滚动到最右边
                            page.evaluate("""
                                const container = document.querySelector('.ag-body-horizontal-scroll-viewport');
                                if (container) {
                                    container.scrollLeft = container.scrollWidth;
                                }
                            """)

                            page.locator(f"//span[@class='cell-renderer-content-text']").first.wait_for(state = "visible",timeout=60000)

                            try:
                                cell_renderer.wait_for(state="visible", timeout=5000)
                            except  Exception as e:

                                pass

                            if cell_renderer.count() > 0:
                                page.wait_for_timeout(1000)
                                text = cell_renderer.first.text_content()
                                print( f"{key}-{product}:{text}")
                                product_dict[operator][key][product] = text
                            else :

                                print(f"{key}-{product}:无")
                                product_dict[operator][key][product] = None



            else:
                print("ip检测不通过，请检查")
        except:
            print("脚本运行异常:" + traceback.format_exc())
        finally:

            print(f"=====关闭店铺：{store_name}=====")
            close_store(store_id)
    print(product_dict)
    return product_dict

def check_verified_status( page):

    continue_shopping_button = page.locator("span.a-button.a-button-primary.a-span12")
    while  continue_shopping_button.is_visible():
        try:
            continue_shopping_button.click(force=True, timeout=5000)
        except:
            pass
        continue_shopping_button = page.locator("span.a-button.a-button-primary.a-span12")
    verify_code = page.frame_locator('iframe[src*="twostep.html"]').get_by_text("验证码获取成功")
    if verify_code.is_visible():
        page.get_by_role("button", name="登录").click(force=True, timeout=5000)


def use_all_browser_run_task_with_thread_pool(browser_list, max_threads=3):
    """
    使用线程池控制最大并发线程数
    :param browser_list: 店铺列表
    :param max_threads: 最大并发线程数
    """
    merged_result = {}
    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        # executor.map 会按顺序返回每个任务的返回值
        result_iterator = executor.map(use_one_browser_run_task, browser_list)
        for result in result_iterator:
            # 深度合并三层嵌套字典：操作人 -> 国家 -> 商品
            for operator, countries_data in result.items():
                if operator not in merged_result:
                    merged_result[operator] = {}

                # 合并该操作人下的所有国家数据
                for country, products_data in countries_data.items():
                    if country not in merged_result[operator]:
                        merged_result[operator][country] = {}

                    # 合并该国家下的所有商品数据
                    merged_result[operator][country].update(products_data)

    print(merged_result)
    return merged_result


from openpyxl import load_workbook
from datetime import datetime, timedelta


def write_result_to_excel(result_dict, file_path):
    """
    将查询结果写入Excel文件，按日期追加新列
    :param result_dict: 操作人->站点->商品 的嵌套字典
    :param file_path: Excel文件路径
    """
    wb = load_workbook(file_path)

    # 获取今天的日期，格式如 5.19
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    date_str = f"{yesterday.year}/{yesterday.month}/{yesterday.day}"

    # 遍历每个操作人（对应sheet页）
    for operator_name, sites_data in result_dict.items():
        # 检查sheet是否存在
        if operator_name not in wb.sheetnames:
            print(f"警告: Sheet '{operator_name}' 不存在，跳过")
            continue

        ws = wb[operator_name]

        # 查找日期列的位置
        # 第一行是标题行，从C列开始是日期列
        last_col = ws.max_column
        new_col = last_col + 1

        # 在第一行写入今天的日期
        ws.cell(row=1, column=new_col, value=date_str)

        # 先处理合并单元格，构建一个行号到站点的映射
        row_to_site = {}
        current_site = None

        # 遍历合并单元格区域
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col == 1 and merged_range.max_col == 1:  # 只处理A列的合并
                # 获取合并区域的值（合并单元格的值在左上角）
                site_value = ws.cell(row=merged_range.min_row, column=1).value
                if site_value:
                    # 为该合并区域的所有行赋值
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        row_to_site[row] = str(site_value)

        # 遍历所有行，填充未合并的站点名称
        for row in range(2, ws.max_row + 1):
            if row not in row_to_site:
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    row_to_site[row] = str(cell_value)
                elif current_site:
                    row_to_site[row] = current_site

        # 遍历站点和商品，找到对应的行并填入数据
        for site_name, products_data in sites_data.items():
            # 处理站点名称匹配（去掉"站"字）
            site_display_name = site_name.replace('站', '')
            if site_name == '美国本土店铺':
                site_display_name = '美国'

            # 遍历所有行，找到匹配的站点和商品
            for row in range(2, ws.max_row + 1):
                # 使用映射的站点名称
                cell_site = row_to_site.get(row, '')

                if cell_site and site_display_name in cell_site:
                    # B列是商品关键词
                    cell_keyword = ws.cell(row=row, column=2).value
                    if cell_keyword and str(cell_keyword).strip() in products_data:
                        # 找到匹配的行，写入数据
                        value = products_data[str(cell_keyword).strip()]
                        ws.cell(row=row, column=new_col, value=value if value else '')

    # 保存文件
    wb.save(file_path)
    print(f"\n=====数据已保存到: {file_path}=====")
    print(f"日期列: {date_str}")



def excute_result(result):
    # 抽离美国本土店铺数据
    normal_result = {}
    us_local_result = {}

    for operator, sites_dict in result.items():
        normal_result[operator] = {}
        us_local_result[operator] = {}

        for site, products in sites_dict.items():
            if site == "美国本土店铺":
                us_local_result[operator][site] = products
            else:
                normal_result[operator][site] = products

    # 清理空的操作人条目
    normal_result = {k: v for k, v in normal_result.items() if v}
    us_local_result = {k: v for k, v in us_local_result.items() if v}

    print("普通数据:", normal_result)
    print("美国本土店铺数据:", us_local_result)

    browser_list= []
    if normal_result:

        browser_dict = {
            "browserName": "巧逗豆-US",
            "browserOauth": "27153125947492",
            "result_dict": normal_result
        }
        browser_list.append(browser_dict)

    if us_local_result:
        browser_dict_us_local = {
            "browserName": "上海天安-US本土店",
            "browserOauth": "27678664116395",
            "result_dict": us_local_result
        }
        browser_list.append(browser_dict_us_local)

    # """打开第一个店铺运行脚本"""
    # normal_result = use_one_browser_run_task(browser_dict,normal_result)
    # us_local_result = use_one_browser_run_task(browser_dict_us_local,us_local_result)




    # """多线程并发打开所有店铺运行脚本，max_threads设置最大线程数"""
    result = use_all_browser_run_task_with_thread_pool(browser_list, max_threads=3)
    return  result

def get_null_items(result_dict):
    """
    提取所有值为None的商品，构建重试字典
    :param result_dict: 原始结果字典
    :return: 只包含None值商品的字典 {操作人: {站点: [商品列表]}}
    """
    null_dict = {}

    for operator, sites_data in result_dict.items():
        for site, products_data in sites_data.items():
            null_products = [product for product, value in products_data.items() if value is None]

            if null_products:
                if operator not in null_dict:
                    null_dict[operator] = {}
                null_dict[operator][site] = null_products

    return null_dict


def merge_null_results(original_dict, retry_dict):
    """
    将重试结果合并到原始结果中（只覆盖None值）
    :param original_dict: 原始结果字典
    :param retry_dict: 重试结果字典
    :return: 合并后的字典
    """
    merged_dict = {}

    # 深拷贝原始字典
    for operator, sites_data in original_dict.items():
        if operator not in merged_dict:
            merged_dict[operator] = {}
        for site, products_data in sites_data.items():
            if site not in merged_dict[operator]:
                merged_dict[operator][site] = {}
            for product, value in products_data.items():
                merged_dict[operator][site][product] = value

    # 只覆盖重试结果中非None的值
    for operator, sites_data in retry_dict.items():
        for site, products_data in sites_data.items():
            if operator not in merged_dict:
                merged_dict[operator] = {}
            if site not in merged_dict[operator]:
                merged_dict[operator][site] = {}

            for product, value in products_data.items():
                # 只有重试后的值不为None时才覆盖
                if value is not None:
                    merged_dict[operator][site][product] = value

    return merged_dict


def excute_result_with_retry(result, max_retry_times=3):
    """
    执行查询并检测空值，如果空值超过3个则重试
    :param original_result: 原始查询字典（操作人->站点->商品列表）
    :param max_retry_times: 最大重试次数
    :return: 最终结果字典
    """


    for retry_num in range(1, max_retry_times + 1):
        null_dict = get_null_items(result)

        # 统计空值总数
        null_count = sum(len(products) for sites in null_dict.values() for products in sites.values())
        print(f"\n===== 空值检测 =====")
        print(f"空值数量: {null_count}")

        if null_count > 0:
            print(f"空值超过3个，开始第{retry_num}次重试...")
            print(f"需要重试的商品: {null_dict}")

            # 重新执行查询（只查询空值部分）
            retry_result = excute_result(null_dict)

            # 合并结果
            result = merge_null_results(result, retry_result)
            print("重试完成，已合并结果")
        else:
            print(f"空值数量({null_count})不超过3个，无需重试")
            break

    return result

def main(args):


    # todo 2、修改用户登录信息，使用企业登录




    # 终止紫鸟客户端已启动的进程
    # todo 3、v5与v6的进程名不同，按版本修改v5或v6
    kill_process(version="v6")

    print("=====启动客户端=====")
    start_browser()
    print("=====更新内核=====")
    update_core()


    get_browser_list()
    file_path = os.getenv("CAMPAIGN_SPEND_EXCEL_PATH")
    if not file_path:
        raise RuntimeError("CAMPAIGN_SPEND_EXCEL_PATH 未设置，无法读取广告花费数据表")
    original_result = read_excel_to_nested_dict(file_path)

    include_null_result = excute_result(original_result)


    result = excute_result_with_retry( include_null_result, max_retry_times=3)
    write_result_to_excel(result, file_path)
    print(result)
    # """关闭客户端"""
    get_exit()


if __name__ == '__main__':
    main([])
