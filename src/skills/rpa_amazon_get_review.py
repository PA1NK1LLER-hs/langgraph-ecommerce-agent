"""Amazon 评论和星级采集技能。

通过 Playwright 多浏览器并发采集 Amazon 商品评论数和星级。
"""

import os
import re
import time
import queue
import threading
import asyncio
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

from openpyxl import Workbook
import pandas as pd
from langchain_core.tools import tool

from pydantic import BaseModel, Field

# ── 配置 ──
amazon_country_dict = {
    "美国": "https://www.amazon.com/",
    "澳洲": "https://www.amazon.com.au/",
    "英国": "https://www.amazon.co.uk/",
    "日本": "https://www.amazon.co.jp/",
    "新加坡": "https://www.amazon.com.sg/",
    "法国": "https://www.amazon.fr/",
    "德国": "https://www.amazon.de/",
    "意大利": "https://www.amazon.it/",
    "西班牙": "https://www.amazon.es/",
    "加拿大": "https://www.amazon.ca/",
    "墨西哥": "https://www.amazon.com.mx/",
    "巴西": "https://www.amazon.com.br/",
}

CONCURRENT_PAGES_PER_BROWSER = 1
counter = 0
counter_lock = threading.Lock()
failed_list: list = []
failed_list_lock = threading.Lock()
success_list: list = []
success_list_lock = threading.Lock()
count = 0

print_queue = queue.Queue()
print_lock = threading.Lock()


def save_list_to_excel(data_list, file_name, output_path=None, custom_date=None):
    """保存数据到 Excel（保留原有逻辑）。"""
    if not data_list:
        return None
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        desktop_path = os.path.join(os.environ["USERPROFILE"], "Desktop")
        output_dir = os.path.join(desktop_path, "共享文件夹", "3.巧逗豆--亚马逊链接星级和评论数日报表")
        output_path = os.path.join(output_dir, file_name)

        today = custom_date or datetime.now().strftime("%Y-%m-%d")
        today_day = today

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        header_font = Font(bold=True)

        operator_groups = {}
        for item in data_list:
            operator = item.get("operator", "未分配")
            if operator not in operator_groups:
                operator_groups[operator] = []
            operator_groups[operator].append(item)

        if os.path.exists(output_path):
            wb = load_workbook(output_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        for operator, items in operator_groups.items():
            sheet_name = operator[:31]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

            existing_products = {}
            existing_product_order = []
            existing_days = []

            max_row = ws.max_row
            if max_row > 1:
                for col in range(3, ws.max_column + 1):
                    day_val = ws.cell(row=1, column=col).value
                    if day_val:
                        existing_days.append(str(day_val))

                for row in range(2, max_row + 1, 2):
                    name = ws.cell(row=row, column=1).value
                    if name:
                        name = str(name)
                        if name not in existing_product_order:
                            existing_product_order.append(name)
                        if name not in existing_products:
                            existing_products[name] = {"star": {}, "review": {}}
                        for col in range(3, ws.max_column + 1):
                            day = ws.cell(row=1, column=col).value
                            if day:
                                day = str(day)
                                star = ws.cell(row=row, column=col).value
                                review = ws.cell(row=row + 1, column=col).value
                                if day not in existing_products[name]["star"]:
                                    existing_products[name]["star"][day] = ""
                                if day not in existing_products[name]["review"]:
                                    existing_products[name]["review"][day] = ""
                                if star is not None:
                                    existing_products[name]["star"][day] = str(star)
                                if review is not None:
                                    existing_products[name]["review"][day] = str(review)

            if today_day not in existing_days:
                existing_days.append(today_day)
                existing_days.sort(key=lambda x: int(x) if x.isdigit() else 0)

            for item in items:
                name = item.get("name", "")
                star = item.get("star", "")
                review = item.get("review", "")
                if name and name not in existing_product_order:
                    existing_product_order.append(name)
                if name and name not in existing_products:
                    existing_products[name] = {"star": {}, "review": {}}
                if name:
                    existing_products[name]["star"][today_day] = star if star else ""
                    existing_products[name]["review"][today_day] = review if review else ""

            ws.delete_rows(1, ws.max_row)

            cell_a1 = ws["A1"]
            cell_a1.value = "产品名称/日期"
            cell_a1.font = header_font
            cell_a1.border = thin_border
            cell_a1.alignment = Alignment(horizontal="center", vertical="center")

            cell_b1 = ws["B1"]
            cell_b1.value = "星级和评论数"
            cell_b1.font = header_font
            cell_b1.border = thin_border
            cell_b1.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, day in enumerate(existing_days, start=3):
                col_letter = (
                    chr(64 + col_idx) if col_idx <= 26
                    else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
                )
                cell = ws[f"{col_letter}1"]
                cell.value = day
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[col_letter].width = 10

            current_row = 2
            for product_name in existing_product_order:
                if product_name not in existing_products:
                    continue
                product_data = existing_products[product_name]
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row + 1, end_column=1)
                cell_name = ws.cell(row=current_row, column=1, value=product_name)
                cell_name.border = thin_border
                cell_name.alignment = Alignment(horizontal="center", vertical="center")

                cell_star_label = ws.cell(row=current_row, column=2, value="star")
                cell_star_label.border = thin_border
                cell_star_label.alignment = Alignment(horizontal="center", vertical="center")

                cell_review_label = ws.cell(row=current_row + 1, column=2, value="reviews")
                cell_review_label.border = thin_border
                cell_review_label.alignment = Alignment(horizontal="center", vertical="center")

                for col_idx, day in enumerate(existing_days, start=3):
                    star = product_data["star"].get(day, "")
                    review = product_data["review"].get(day, "")
                    col_letter = (
                        chr(64 + col_idx) if col_idx <= 26
                        else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
                    )
                    cell_star = ws[f"{col_letter}{current_row}"]
                    cell_star.value = star
                    cell_star.border = thin_border
                    cell_star.alignment = Alignment(horizontal="center", vertical="center")

                    cell_review = ws[f"{col_letter}{current_row + 1}"]
                    cell_review.value = review
                    cell_review.border = thin_border
                    cell_review.alignment = Alignment(horizontal="center", vertical="center")

                ws.row_dimensions[current_row].height = 20
                ws.row_dimensions[current_row + 1].height = 20
                current_row += 2

            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 15

        wb.save(output_path)
        return output_path
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None


async def clear_browser_cache(context, page):
    try:
        await context.clear_cookies()
        try:
            await page.evaluate("""() => {
                localStorage.clear();
                sessionStorage.clear();
                if (window.indexedDB) {
                    indexedDB.databases().then(dbs => {
                        dbs.forEach(db => indexedDB.deleteDatabase(db.name));
                    });
                }
                if ('caches' in window) {
                    caches.keys().then(names => {
                        names.forEach(name => caches.delete(name));
                    });
                }
            }""")
        except Exception:
            pass
        await page.reload(wait_until="domcontentloaded")
    except Exception as e:
        pass


async def check_verified_status(context, page):
    dogs_of_amazon_link = page.locator('a[href="/dogsofamazon"] img[id="d"]')
    wrong = page.get_by_text("处理您的请求时出错")
    if await dogs_of_amazon_link.is_visible() or await wrong.is_visible():
        await clear_browser_cache(context, page)
        return True
    continue_shopping_button = page.locator("span.a-button.a-button-primary.a-span12")
    while await continue_shopping_button.is_visible():
        try:
            await continue_shopping_button.click(force=True, timeout=5000)
        except Exception:
            pass
        continue_shopping_button = page.locator("span.a-button.a-button-primary.a-span12")


async def process_single_asin(asin_dict, browser, semaphore):
    from playwright_stealth.stealth import Stealth
    stealth = Stealth()

    async with semaphore:
        browser_context = await browser.new_context()
        page = await browser_context.new_page()

        asin = asin_dict["asin"]
        country = asin_dict["country"]
        url = f"{amazon_country_dict[country]}dp/{asin}"

        await stealth.apply_stealth_async(page)
        star = None
        review = None
        try:
            await page.goto(url, timeout=60000, wait_until="domcontentloaded")
            await check_verified_status(browser_context, page)
            await page.wait_for_load_state(state="domcontentloaded")

            stars = page.locator("#acrPopover > span > a > span").first
            reviews = page.locator("#acrCustomerReviewText").first
            await page.wait_for_selector("#acrPopover", state="visible", timeout=30000)

            if await stars.count() > 0:
                star = await stars.text_content()
            if await reviews.count() > 0:
                review = await reviews.text_content()
                review = re.sub(r"\D", "", review)

            with counter_lock:
                global counter
                counter += 1
                current_counter = counter

            if review or star:
                with success_list_lock:
                    success_list.append({
                        "operator": asin_dict["operator"],
                        "name": asin_dict["name"],
                        "asin": asin,
                        "star": star,
                        "review": review,
                        "country": country,
                        "isSuccess": "成功",
                    })
            else:
                await page.screenshot(path=f"失败/{asin}.png")
                with failed_list_lock:
                    failed_list.append({
                        "operator": asin_dict["operator"],
                        "name": asin_dict["name"],
                        "asin": asin,
                        "star": star,
                        "review": review,
                        "country": country,
                        "isSuccess": "失败",
                    })
        except Exception:
            with failed_list_lock:
                failed_list.append({
                    "operator": asin_dict["operator"],
                    "name": asin_dict["name"],
                    "asin": asin,
                    "star": star,
                    "review": review,
                    "country": country,
                    "isSuccess": "失败",
                })
        finally:
            await page.close()
            await browser_context.close()


async def process_batch_in_browser(asin_batch, batch_num, headless):
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--incognito",
                "--disable-gpu",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-images",
            ],
        )
        semaphore = asyncio.Semaphore(CONCURRENT_PAGES_PER_BROWSER)
        tasks = [process_single_asin(asin_dict, browser, semaphore) for asin_dict in asin_batch]
        await asyncio.gather(*tasks)
        await browser.close()


def run_task(asin_batch, batch_num, headless):
    asyncio.run(process_batch_in_browser(asin_batch, batch_num, headless))


def read_asin_from_excel(excel_path):
    df = pd.read_excel(excel_path)
    data_list = []
    for _, row in df.iterrows():
        name = str(row["品名"]).strip() if pd.notna(row["品名"]) else ""
        asin = str(row["ASIN"]).strip() if pd.notna(row["ASIN"]) else ""
        country = str(row["国家"]).strip() if pd.notna(row["国家"]) else ""
        operator = str(row["负责人"]).strip() if pd.notna(row["负责人"]) else ""
        if asin and country and country in ("美国", "加拿大"):
            data_list.append({"asin": asin, "country": country, "name": name, "operator": operator})
    return data_list


def excute_asin(data_list, headless=True, max_workers=6):
    global counter, count
    counter = 0
    count = len(data_list)
    if count == 0:
        return

    batch_size = (count + max_workers - 1) // max_workers
    batches = [data_list[i:i + batch_size] for i in range(0, count, batch_size)]

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(run_task, batch, i + 1, headless) for i, batch in enumerate(batches)]
        for f in futures:
            f.result()


def calculate_optimal_concurrent():
    import multiprocessing
    logical_cores = multiprocessing.cpu_count()
    physical_cores = logical_cores // 2
    optimal_threads = int(physical_cores * 1.2)
    optimal_threads = max(3, min(optimal_threads, 8))

    if logical_cores <= 4:
        coroutines_per_browser = 2
    elif logical_cores <= 8:
        coroutines_per_browser = 3
    else:
        coroutines_per_browser = 3

    return optimal_threads, coroutines_per_browser


class AmazonGetReviewArgs(BaseModel):
    excel_path: str = Field(
        default="",
        description="Excel 文件路径。为空则使用默认路径。",
    )


@tool(args_schema=AmazonGetReviewArgs)
def amazon_get_review(excel_path: str = "") -> dict:
    """采集 Amazon 商品评论数和星级。从 Excel 读取 ASIN 列表，通过多浏览器并发访问 Amazon 各站点，
采集每个商品的评分（star）和评论数（review），自动重试失败项，结果保存到 Excel 报表。

当用户要求采集 Amazon 评论、获取商品星级和评论数、做 ASIN 报表时使用。"""
    global CONCURRENT_PAGES_PER_BROWSER, failed_list, success_list
    failed_list = []
    success_list = []

    download_path = excel_path or r"\\192.168.10.27\共享文件夹\运营端提供的原始数据\各站点ASIN-评论星级用.xlsx"

    optimal_workers, CONCURRENT_PAGES_PER_BROWSER = calculate_optimal_concurrent()

    data_list = read_asin_from_excel(download_path)
    excute_asin(data_list, headless=True, max_workers=optimal_workers)

    t = 0
    while (len(failed_list) > 5 or t == 0) and len(failed_list) > 0 and t < 10:
        t += 1
        temp_list = failed_list.copy()
        failed_list.clear()
        excute_asin(temp_list, headless=True, max_workers=optimal_workers)

    success_list.extend(failed_list)

    country_groups = {}
    for item in success_list:
        country = item.get("country", "未知国家")
        if country not in country_groups:
            country_groups[country] = []
        country_groups[country].append(item)

    return {"status": "success", "message": country_groups}


if __name__ == "__main__":
    result = amazon_get_review.invoke({"excel_path": ""})
    print(result)
