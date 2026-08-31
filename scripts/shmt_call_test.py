# -*- coding: utf-8 -*-
"""agent 调用世贸通 skill 的只读验证 — 触发工具调用但显式 deny（不执行，不碰业务）。

验证目标：
  1. agent 能正确识别/绑定 tool_shimaotong_submit
  2. agent 收到报关请求后会调用该工具（而非别的方法）
  3. 因 tool_shimaotong_ 是 high 风险，调用会先发 approval_required 事件（此时工具未执行）
测试在收到审批后显式发送 deny → 工具不执行 → 世贸通不会产生任何真实订单。

缓存免疫：每次运行生成带时间戳的唯一假 Excel（路径进 prompt → user_text 不同），
避免命中后端语义缓存（缓存只以 user_text 为键，且会缓存"已拒绝"的最终答复）。
"""

import asyncio
import json
import sys
import time
import uuid

import httpx
import websockets
from openpyxl import Workbook

BASE = "http://127.0.0.1:8080"
WS_URL = "ws://127.0.0.1:8080/ws/chat"
CLAUDE = ("claude_verify", "ClaudeVerify123!")

EXCEL_DIR = r"D:\PycharmProjects\langgraph-agent\workspace"  # MCP 允许目录


def make_fake_excel(path: str) -> None:
    """生成结构真实但 SKU 全空、编号不含 "PO" 的假报关 Excel。

    即使审批意外放行、工具真执行，detect_orders 也识别不到订单 → 报"未找到订单"退出，零写操作。
    """
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    ws = wb.create_sheet("明细6")
    ws.append([
        "编号", "代理", "国家", "目的港", "起运地", "出口抬头",
        "预估海运费", "保险费", "投保金额", "提单抬头人",
        "进仓单号", "报关单号", "FBA编号", "仓库",
        "外箱SKU", "店铺SKU", "品名", "华飞系统型号SKU",
        "合计数量", "数量", "单价", "运费", "价格合计",
        "报关出口单价", "报关出口价格合计",
    ])
    ws.append([
        "TESTFAKE20260828-001", "测试代理", "美国", "纽约", "上海", "测试抬头",
        "1000", "10", "50000", "测试提单人",
        "", "", "", "测试仓库",
        "", "", "", "",  # 外箱/店铺/品名/华飞SKU 全空
        "", "", "", "", "", "",
    ])
    wb.save(path)


async def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    # 唯一假 Excel + 唯一 prompt → 缓存免疫
    fake_name = f"shmt_test_{time.strftime('%H%M%S')}_{uuid.uuid4().hex[:6]}.xlsx"
    fake_excel = f"{EXCEL_DIR}\\{fake_name}"
    make_fake_excel(fake_excel)
    print(f"[setup] 假 Excel: {fake_excel}")

    try:
        print(f"=== agent 调用世贸通 skill 验证（显式 deny，工具不执行）===")
        r = httpx.post(f"{BASE}/api/auth/login",
                       json={"username": CLAUDE[0], "password": CLAUDE[1]},
                       headers={"X-Real-IP": "10.7.2.2"})
        token = r.json()["access_token"]
        print(f"登录成功 token 长度={len(token)}")

        async with websockets.connect(WS_URL, open_timeout=15) as ws:
            await ws.send(json.dumps({"type": "auth", "token": token}))
            for _ in range(4):
                evt = json.loads(await asyncio.wait_for(ws.recv(), timeout=15))
                if evt.get("type") == "thread_id":
                    print(f"thread_id: {evt.get('thread_id')}")
                    break

            prompt = (
                f"请调用世贸通抬头报关工具，处理报关 Excel 文件：{fake_excel}。"
                "只保存订单，不要提交报关资料。"
            )
            print(f"\n提示词: {prompt[:70]}...")
            await ws.send(json.dumps({"message": prompt}))

            approval_seen = None
            tool_call_seen = None
            done = False
            t0 = asyncio.get_event_loop().time()
            while asyncio.get_event_loop().time() - t0 < 90:
                try:
                    evt = json.loads(await asyncio.wait_for(ws.recv(), timeout=90))
                except Exception:
                    break
                ty = evt.get("type")
                content = evt.get("content", "") or evt.get("tool", "") or evt.get("message", "")
                print(f"  [{ty}] {str(content)[:120]}")
                if ty == "approval_required":
                    approval_seen = evt
                    # 显式 deny：干净完成，工具不执行
                    await ws.send(json.dumps({"type": "approval_decision", "decision": "deny"}))
                    break
                elif ty == "tool_call":
                    tool_call_seen = evt.get("tool")
                elif ty == "done":
                    done = True
                elif ty == "error":
                    break

        print("\n" + "=" * 60)
        ok = approval_seen is not None
        if approval_seen:
            tools = [tc.get("name") for tc in (approval_seen.get("calls") or [])]
            print(f"[{'PASS' if ok else 'FAIL'}] agent 调用了工具，触发审批:")
            print(f"  审批工具列表: {tools}")
            print(f"  是否含 tool_shimaotong_submit: {'tool_shimaotong_submit' in tools}")
            ok = ok and "tool_shimaotong_submit" in tools
        else:
            print(f"[FAIL] 未收到 approval_required。tool_call 捕捉到: {tool_call_seen}, done={done}")
    finally:
        import os
        try:
            os.remove(fake_excel)
            print(f"[cleanup] 已删除假 Excel: {fake_name}")
        except Exception:
            pass

    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    asyncio.run(main())
